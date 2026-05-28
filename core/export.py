"""
export.py — Excel and CSV export for the Cola Market Simulation.
Produces a multi-sheet .xlsx workbook and a flat .csv summary.
"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from core.model import (
    MarketConditions,
    EquilibriumResult,
    Financials,
    price_sensitivity,
)

# ---------------------------------------------------------------------------
# Colour palette (openpyxl uses ARGB hex, no leading #)
# ---------------------------------------------------------------------------
C_NAVY    = "FF1B3A6B"
C_TEAL    = "FF0D7377"
C_LIGHT   = "FFE8F4F8"
C_WHITE   = "FFFFFFFF"
C_BLACK   = "FF000000"
C_RED_BG  = "FFFDE8E8"
C_GREEN_BG= "FFE8F8EE"
C_YELLOW  = "FFFFF2CC"
C_BLUE_IN = "FF0000FF"   # blue = hardcoded input (industry standard)
C_GREY_HDR= "FFD9E1F2"
C_BORDER  = "FFB0C4DE"

def _thin_border(sides="all"):
    thin = Side(style="thin", color=C_BORDER)
    none = Side(style=None)
    if sides == "all":
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    if sides == "bottom":
        return Border(bottom=thin)
    if sides == "top_bottom":
        return Border(top=thin, bottom=thin)
    return Border()

def _hdr_font(sz=11, bold=True, color=C_WHITE):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def _cell_font(sz=10, bold=False, color=C_BLACK, italic=False):
    return Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)

def _set(ws, row, col, value, font=None, fill=None, align=None,
         border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    if num_fmt: c.number_format = num_fmt
    return c

def _merge_hdr(ws, row, c1, c2, text, bg=C_NAVY, fg=C_WHITE, sz=12):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row,   end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font      = Font(name="Arial", size=sz, bold=True, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell

# ---------------------------------------------------------------------------
# Sheet 1 — Summary
# ---------------------------------------------------------------------------

def _sheet_summary(wb, mc: MarketConditions,
                   eq: EquilibriumResult, fin: Financials,
                   scenario: str):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.row_dimensions[1].height = 28

    # Title bar
    _merge_hdr(ws, 1, 1, 3,
               f"Cola Retail Market Simulation  |  {scenario}",
               bg=C_NAVY, sz=13)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    ts = ws.cell(row=2, column=1,
                 value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ts.font      = _cell_font(sz=9, italic=True, color="FF555555")
    ts.alignment = Alignment(horizontal="center")

    row = 4
    brands = {0: "Independent", 1: "Regional chain", 2: "National chain"}

    sections = [
        ("EQUILIBRIUM SOLUTION", [
            ("Equilibrium price",    f"${eq.price_eq:.4f}",   "per unit"),
            ("Equilibrium quantity", f"{eq.quantity_eq:,.0f}", "units / week"),
            ("Quantity demanded",    f"{eq.quantity_demanded:,.0f}", "units / week"),
            ("Quantity supplied",    f"{eq.quantity_supplied:,.0f}", "units / week"),
            ("Residual excess",      f"{eq.excess:.4f}",      "units"),
            ("Solver method",        eq.method,               ""),
            ("Converged",            "YES" if eq.converged else "NO", ""),
        ]),
        ("WEEKLY FINANCIALS", [
            ("Revenue",              f"${fin.revenue:,.2f}",    ""),
            ("COGS",                 f"${fin.cogs:,.2f}",       ""),
            ("Gross profit",         f"${fin.gross_profit:,.2f}",
             f"({fin.gross_margin_pct:.1f}% margin)"),
            ("Advertising expense",  f"${fin.ad_expense:,.2f}", ""),
            ("Transport expense",    f"${fin.transport_expense:,.2f}", ""),
            ("Fixed overhead",       f"${fin.fixed_overhead:,.2f}", ""),
            ("Total OpEx",           f"${fin.total_opex:,.2f}", ""),
            ("EBIT",                 f"${fin.ebit:,.2f}",       ""),
            ("Tax",                  f"${fin.tax_amount:,.2f}", f"({mc.tax_rate_pct:.1f}%)"),
            ("Net profit",           f"${fin.net_profit:,.2f}",
             f"({fin.net_margin_pct:.1f}% margin)"),
            ("Contribution margin",  f"${fin.contribution_margin:.4f}", "per unit"),
            ("Break-even volume",
             f"{fin.break_even_units:,.0f}" if fin.break_even_units != float("inf") else "N/A",
             "units / week"),
            ("Profitable?",
             "YES" if eq.quantity_eq >= fin.break_even_units else "NO", ""),
        ]),
        ("MARKET CONDITIONS (INPUTS)", [
            ("Own price (seed)",     f"${mc.own_price:.2f}",   "per unit"),
            ("Competitor price",     f"${mc.competitor_price:.2f}", "per unit"),
            ("Wholesale cost",       f"${mc.wholesale_cost:.2f}", "per unit"),
            ("Ad spend",             f"${mc.ad_spend_k:.1f}K", "per week"),
            ("Transport cost",       f"${mc.transport_cost_k:.1f}K", "per week"),
            ("Fixed overhead",       f"${mc.fixed_overhead_k:.1f}K", "per week"),
            ("Tax rate",             f"{mc.tax_rate_pct:.1f}%", ""),
            ("Local income (median)",f"${mc.local_income_k:.0f}K", "per year"),
            ("Unemployment rate",    f"{mc.unemployment_pct:.1f}%", ""),
            ("Population density",   f"{mc.pop_density:,.0f}", "people / sq mi"),
            ("Consumer satisfaction",f"{mc.consumer_sat:.1f}/10", ""),
            ("Employee satisfaction",f"{mc.employee_sat:.1f}/10", ""),
            ("Upstream market power",f"{mc.upstream_power:.2f}", "0=none 1=monopoly"),
            ("Season index",         f"{mc.season_index:.2f}",  "1.0=average"),
            ("Temperature",          f"{mc.temperature_f:.0f}°F", ""),
            ("Shopper age index",    f"{mc.age_index:.1f} yrs", ""),
            ("Health trend",         f"{mc.health_trend:+.2f}", "0=neutral"),
            ("In-store promotion",   "Yes" if mc.promo_dummy else "No", ""),
            ("Brand type",           brands.get(mc.brand, "Unknown"), ""),
            ("Store count",          f"{mc.store_count:.0f}", ""),
            ("Energy cost index",    f"{mc.energy_cost_idx:.2f}", "1.0=baseline"),
            ("Input scarcity",       f"{mc.input_scarcity:.2f}", "0=none 1=severe"),
            ("Regulatory burden",    f"{mc.regulatory_burden:.2f}", "0=none 1=max"),
            ("Capacity utilisation", f"{mc.capacity_util_pct:.1f}%", ""),
            ("Time since baseline",  f"{mc.time_months:.0f} months", ""),
        ]),
    ]

    for sec_title, items in sections:
        # Section header
        _merge_hdr(ws, row, 1, 3, sec_title, bg=C_TEAL, sz=11)
        row += 1
        # Column sub-headers
        for col, txt in enumerate(["Parameter", "Value", "Unit / Note"], 1):
            _set(ws, row, col, txt,
                 font=_hdr_font(sz=10, color=C_BLACK),
                 fill=PatternFill("solid", fgColor=C_GREY_HDR),
                 align=Alignment(horizontal="center"),
                 border=_thin_border())
        row += 1
        for i, (label, val, note) in enumerate(items):
            bg = C_LIGHT if i % 2 == 0 else C_WHITE
            fill = PatternFill("solid", fgColor=bg[2:])  # strip FF prefix
            # label
            _set(ws, row, 1, label,
                 font=_cell_font(sz=10),
                 fill=PatternFill("solid", fgColor=bg[2:]),
                 border=_thin_border("bottom"))
            # value
            profit_row = (label == "Net profit")
            val_color = ("FF006400" if fin.net_profit >= 0 else "FFCC0000") \
                        if profit_row else C_BLUE_IN
            _set(ws, row, 2, val,
                 font=Font(name="Arial", size=10, bold=profit_row,
                           color=val_color),
                 fill=PatternFill("solid", fgColor=bg[2:]),
                 align=Alignment(horizontal="right"),
                 border=_thin_border("bottom"))
            # note
            _set(ws, row, 3, note,
                 font=_cell_font(sz=9, italic=True, color="FF666666"),
                 fill=PatternFill("solid", fgColor=bg[2:]),
                 border=_thin_border("bottom"))
            row += 1
        row += 1  # gap between sections

# ---------------------------------------------------------------------------
# Sheet 2 — Price Sensitivity
# ---------------------------------------------------------------------------

def _sheet_sensitivity(wb, mc: MarketConditions, eq: EquilibriumResult):
    ws = wb.create_sheet("Price Sensitivity")
    ws.sheet_view.showGridLines = False

    cols = {"Price ($/unit)": 14, "Qd (units/wk)": 16,
            "Qs (units/wk)": 16, "Excess Demand": 16,
            "Revenue ($)": 16, "Net Profit ($)": 16, "Near Eq?": 10}
    for i, (name, width) in enumerate(cols.items(), 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    _merge_hdr(ws, 1, 1, len(cols),
               "Price Sensitivity Analysis  —  $0.50 to $4.00/unit",
               bg=C_NAVY, sz=12)
    ws.row_dimensions[1].height = 24

    for i, name in enumerate(cols.keys(), 1):
        _set(ws, 2, i, name,
             font=_hdr_font(sz=10, color=C_BLACK),
             fill=PatternFill("solid", fgColor=C_GREY_HDR[2:]),
             align=Alignment(horizontal="center"),
             border=_thin_border())

    rows = price_sensitivity(mc, steps=30)
    for r_idx, r in enumerate(rows, 3):
        is_eq = abs(r["price"] - eq.price_eq) < 0.05
        bg = "FFF0FFF0" if is_eq else ("FFECF4FB" if r_idx % 2 == 0 else C_WHITE[2:])
        rev  = r["price"] * r["Qs"]
        opex = (mc.ad_spend_k + mc.transport_cost_k + mc.fixed_overhead_k) * 1000
        profit = (rev - mc.wholesale_cost * r["Qs"] - opex) * (1 - mc.tax_rate_pct / 100)

        vals = [r["price"], r["Qd"], r["Qs"], r["excess"], rev, profit,
                "← EQ" if is_eq else ""]
        fmts = ["$#,##0.00", "#,##0", "#,##0", "+#,##0;-#,##0;0",
                "$#,##0", "$#,##0", "@"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            profit_col = ci == 6
            clr = ("FF006400" if profit >= 0 else "FFCC0000") if profit_col else C_BLACK[2:]
            _set(ws, r_idx, ci, v,
                 font=Font(name="Arial", size=10, color=clr,
                           bold=is_eq),
                 fill=PatternFill("solid", fgColor=bg),
                 align=Alignment(horizontal="right" if ci < 7 else "center"),
                 num_fmt=fmt,
                 border=_thin_border("bottom"))

# ---------------------------------------------------------------------------
# Sheet 3 — Demand / Supply breakdown
# ---------------------------------------------------------------------------

def _sheet_factors(wb, mc: MarketConditions, eq: EquilibriumResult):
    ws = wb.create_sheet("Factor Breakdown")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20

    _merge_hdr(ws, 1, 1, 4, "Demand & Supply Factor Breakdown at Equilibrium",
               bg=C_NAVY, sz=12)
    ws.row_dimensions[1].height = 24

    import math
    p = eq.price_eq

    demand_factors = [
        ("Own price  [log-lin]",         -1.60 * math.log(p)),
        ("Competitor price  [log-lin]",   0.90 * math.log(max(mc.competitor_price, 0.01))),
        ("Ad spend  [log-lin]",           0.25 * math.log(max(mc.ad_spend_k, 0.01))),
        ("Income  [log-lin]",             0.50 * math.log(max(mc.local_income_k, 0.01))),
        ("Unemployment  [linear]",       -0.030 * mc.unemployment_pct),
        ("Consumer sat (linear)  [quad]", 0.18  * mc.consumer_sat),
        ("Consumer sat (quad)  [quad]",  -0.012 * mc.consumer_sat**2),
        ("Time (linear)  [cubic]",        0.025 * mc.time_months),
        ("Time (quad)  [cubic]",         -0.0008 * mc.time_months**2),
        ("Time (cubic)  [cubic]",         4e-6  * mc.time_months**3),
        ("Season  [log-lin]",             0.40  * math.log(max(mc.season_index, 0.01))),
        ("Brand effect  [dummy]",         [0.0, 0.15, 0.35][mc.brand]),
        ("Population density  [log-lin]", 0.30  * math.log(max(mc.pop_density, 1))),
        ("Age (linear)  [quad]",          0.040 * mc.age_index),
        ("Age (quad)  [quad]",           -0.0007 * mc.age_index**2),
        ("Health trend  [linear]",       -0.20  * mc.health_trend),
        ("Promotion  [linear]",           0.18  * mc.promo_dummy),
        ("Temperature (linear)  [quad]",  0.006  * mc.temperature_f),
        ("Temperature (quad)  [quad]",   -3.5e-5 * mc.temperature_f**2),
        ("Intercept  a₀",                 2.791),
    ]

    supply_factors = [
        ("Own price  [log-lin]",          1.20 * math.log(p)),
        ("Wholesale cost  [log-lin]",    -1.40 * math.log(max(mc.wholesale_cost, 0.01))),
        ("Tax rate  [linear]",           -0.015 * mc.tax_rate_pct),
        ("Transport cost  [log-lin]",    -0.08  * math.log(max(mc.transport_cost_k, 0.01))),
        ("Upstream power (lin)  [cubic]",-0.50  * mc.upstream_power),
        ("Upstream power (quad) [cubic]",-0.30  * mc.upstream_power**2),
        ("Upstream power (cub)  [cubic]", 0.10  * mc.upstream_power**3),
        ("Employee sat (lin)  [quad]",    0.14  * mc.employee_sat),
        ("Employee sat (quad) [quad]",   -0.008 * mc.employee_sat**2),
        ("Capacity utilisation  [linear]",0.005 * mc.capacity_util_pct),
        ("Energy cost  [log-lin]",       -0.35  * math.log(max(mc.energy_cost_idx, 0.01))),
        ("Input scarcity  [linear]",     -0.40  * mc.input_scarcity),
        ("Regulatory burden  [linear]",  -0.25  * mc.regulatory_burden),
        ("Store count  [log-lin]",        0.20  * math.log(max(mc.store_count, 1))),
        ("Intercept  b₀",                 7.513),
    ]

    for sec_row, (title, factors) in enumerate(
            [("DEMAND FACTORS  →  contribution to ln(Qd)", demand_factors),
             ("SUPPLY FACTORS  →  contribution to ln(Qs)", supply_factors)]):
        base_row = 3 + sec_row * (len(demand_factors) + 5)
        _merge_hdr(ws, base_row, 1, 4, title, bg=C_TEAL, sz=11)
        for ci, hdr in enumerate(["Factor", "Contribution", "Form", "Direction"], 1):
            _set(ws, base_row + 1, ci, hdr,
                 font=_hdr_font(sz=10, color=C_BLACK),
                 fill=PatternFill("solid", fgColor=C_GREY_HDR[2:]),
                 align=Alignment(horizontal="center"),
                 border=_thin_border())
        for i, (fname, contrib) in enumerate(factors):
            r = base_row + 2 + i
            bg = "FFECF4FB" if i % 2 == 0 else C_WHITE[2:]
            form = fname.split("[")[-1].rstrip("]") if "[" in fname else "constant"
            direction = "positive ▲" if contrib >= 0 else "negative ▼"
            dir_color = "FF006400" if contrib >= 0 else "FFCC0000"
            for ci, (v, fmt, clr) in enumerate([
                (fname.split("  [")[0], None, C_BLACK[2:]),
                (contrib, "0.0000", "FF0000FF"),
                (form, None, "FF555555"),
                (direction, None, dir_color),
            ], 1):
                _set(ws, r, ci, v,
                     font=Font(name="Arial", size=10, color=clr),
                     fill=PatternFill("solid", fgColor=bg),
                     align=Alignment(horizontal="right" if ci == 2 else "left"),
                     num_fmt=fmt,
                     border=_thin_border("bottom"))

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_xlsx(mc: MarketConditions, eq: EquilibriumResult,
                fin: Financials, scenario: str = "Simulation") -> bytes:
    """Build the workbook and return raw bytes (for Streamlit download or disk)."""
    wb = openpyxl.Workbook()
    _sheet_summary(wb, mc, eq, fin, scenario)
    _sheet_sensitivity(wb, mc, eq)
    _sheet_factors(wb, mc, eq)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_xlsx(mc: MarketConditions, eq: EquilibriumResult,
              fin: Financials, path: str, scenario: str = "Simulation") -> None:
    """Write the workbook directly to a file path."""
    data = export_xlsx(mc, eq, fin, scenario)
    with open(path, "wb") as f:
        f.write(data)


def export_csv(mc: MarketConditions, eq: EquilibriumResult,
               fin: Financials, scenario: str = "Simulation") -> str:
    """Return a flat CSV string of the full simulation result."""
    import csv, io as _io
    buf = _io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Cola Market Simulation Export"])
    w.writerow(["Scenario", scenario])
    w.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    w.writerow([])

    w.writerow(["=== EQUILIBRIUM ==="])
    w.writerow(["equilibrium_price",    round(eq.price_eq, 6)])
    w.writerow(["equilibrium_quantity", round(eq.quantity_eq, 2)])
    w.writerow(["quantity_demanded",    round(eq.quantity_demanded, 2)])
    w.writerow(["quantity_supplied",    round(eq.quantity_supplied, 2)])
    w.writerow(["residual_excess",      round(eq.excess, 6)])
    w.writerow(["converged",            eq.converged])
    w.writerow(["solver_method",        eq.method])
    w.writerow([])

    w.writerow(["=== FINANCIALS (weekly) ==="])
    for k, v in [
        ("revenue",            fin.revenue),
        ("cogs",               fin.cogs),
        ("gross_profit",       fin.gross_profit),
        ("gross_margin_pct",   fin.gross_margin_pct),
        ("ad_expense",         fin.ad_expense),
        ("transport_expense",  fin.transport_expense),
        ("fixed_overhead",     fin.fixed_overhead),
        ("total_opex",         fin.total_opex),
        ("ebit",               fin.ebit),
        ("tax_amount",         fin.tax_amount),
        ("net_profit",         fin.net_profit),
        ("net_margin_pct",     fin.net_margin_pct),
        ("contribution_margin",fin.contribution_margin),
        ("break_even_units",   fin.break_even_units),
    ]:
        w.writerow([k, round(v, 4)])
    w.writerow([])

    w.writerow(["=== MARKET CONDITIONS ==="])
    for k, v in mc.__dict__.items():
        w.writerow([k, v])
    w.writerow([])

    w.writerow(["=== PRICE SENSITIVITY ==="])
    w.writerow(["price", "qd", "qs", "excess", "profit"])
    for r in price_sensitivity(mc, steps=30):
        rev  = r["price"] * r["Qs"]
        opex = (mc.ad_spend_k + mc.transport_cost_k + mc.fixed_overhead_k) * 1000
        profit = (rev - mc.wholesale_cost * r["Qs"] - opex) * (1 - mc.tax_rate_pct / 100)
        w.writerow([round(r["price"], 2), round(r["Qd"], 1),
                    round(r["Qs"], 1), round(r["excess"], 1), round(profit, 2)])

    return buf.getvalue()
